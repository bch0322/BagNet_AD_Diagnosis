import nibabel as nib
from matplotlib import gridspec
import utils as ut
import numbers
from torch.nn import functional as F
from torch.autograd import Variable
import math
import torch
import torch.nn as nn
import shutil
import numpy as np
import os
import setting as st
import setting_2 as fst
import nibabel as nib
import matplotlib
matplotlib.use('Agg')
import torch
from sklearn import metrics
from sklearn.metrics import r2_score, mean_squared_error
from sklearn.metrics import confusion_matrix
from sklearn.utils.multiclass import unique_labels

import matplotlib.pyplot as plt
import utils
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from collections import deque

class FocalLoss(nn.Module):
    def __init__(self, gamma=0, alpha=None, size_average=True):
        super(FocalLoss, self).__init__()
        self.gamma = gamma
        self.alpha = alpha
        # if isinstance(alpha,(float,int,long)): self.alpha = torch.Tensor([alpha,1-alpha])
        if isinstance(alpha, (float, int)): self.alpha = torch.Tensor([alpha, 1 - alpha])
        if isinstance(alpha,list): self.alpha = torch.Tensor(alpha)
        self.size_average = size_average

    def forward(self, input, target):
        """
        :param input: [batch, 2]
        :param target: [batch, 1]
        :return:
        """
        if input.dim()>2:
            input = input.view(input.size(0),input.size(1),-1)  # N,C,H,W => N,C,H*W
            input = input.transpose(1,2)    # N,C,H*W => N,H*W,C
            input = input.contiguous().view(-1,input.size(2))   # N,H*W,C => N*H*W,C
        target = target.view(-1,1)

        logpt = F.log_softmax(input, dim=1)
        logpt = logpt.gather(1,target)
        logpt = logpt.view(-1)
        pt = Variable(logpt.data.exp())

        if self.alpha is not None:
            if self.alpha.type()!=input.data.type():
                self.alpha = self.alpha.type_as(input.data)
            at = self.alpha.gather(0,target.data.view(-1))
            logpt = logpt * Variable(at)

        loss = -1 * (1-pt)**self.gamma * logpt
        if self.size_average: return loss.mean()
        else: return loss.sum()


def get_input_optimizer(input_img):
    optimizer = torch.optim.Adam([input_img.requires_grad_()], lr=0.001)
    return optimizer


class GaussianSmoothing(nn.Module):
    """
    Apply gaussian smoothing on a
    1d, 2d or 3d tensor. Filtering is performed seperately for each channel
    in the input using a depthwise convolution.
    Arguments:
        channels (int, sequence): Number of channels of the input tensors. Output will
            have this number of channels as well.
        kernel_size (int, sequence): Size of the gaussian kernel.
        sigma (float, sequence): Standard deviation of the gaussian kernel.
        dim (int, optional): The number of dimensions of the data.
            Default value is 2 (spatial).
    """
    def __init__(self, channels, kernel_size, sigma, dim=3):
        super(GaussianSmoothing, self).__init__()
        if isinstance(kernel_size, numbers.Number):
            kernel_size = [kernel_size] * dim
        if isinstance(sigma, numbers.Number):
            sigma = [sigma] * dim

        # The gaussian kernel is the product of the
        # gaussian function of each dimension.
        kernel = 1
        meshgrids = torch.meshgrid(
            [
                torch.arange(size, dtype=torch.float32)
                for size in kernel_size
            ]
        )
        for size, std, mgrid in zip(kernel_size, sigma, meshgrids):
            mean = (size - 1) / 2
            kernel *= 1 / (std * math.sqrt(2 * math.pi)) * \
                      torch.exp(-((mgrid - mean) / std) ** 2 / 2)

        # Make sure sum of values in gaussian kernel equals 1.
        kernel = kernel / torch.sum(kernel)

        # Reshape to depthwise convolutional weight
        kernel = kernel.view(1, 1, *kernel.size())
        kernel = kernel.repeat(channels, *[1] * (kernel.dim() - 1))
        kernel = kernel.cuda()
        self.register_buffer('weight', kernel)
        self.groups = channels
        self.kernel_size = kernel_size
        if dim == 1:
            self.conv = F.conv1d
        elif dim == 2:
            self.conv = F.conv2d
        elif dim == 3:
            self.conv = F.conv3d
        else:
            raise RuntimeError(
                'Only 1, 2 and 3 dimensions are supported. Received {}.'.format(dim)
            )

    def forward(self, input):
        """
        Apply gaussian filter to input.
        Arguments:
            input (torch.Tensor): Input to apply gaussian filter on.
        Returns:
            filtered (torch.Tensor): Filtered output.
        """
        input = nn.ConstantPad3d(self.kernel_size[0]//2, 0)(input)
        return self.conv(input, weight=self.weight, groups=self.groups)

def model_save_through_validation(fold, epoch, EMS, selected_EMS, ES, model, dir_save_model, metric_1 = 'val_loss', metric_2=None, save_flag = False):
    """ save the model """
    start_eval_epoch = st.early_stopping_start_epoch
    # start_eval_epoch = 1

    tmp_flag = False
    if save_flag == False:
        if epoch >= start_eval_epoch:
            if metric_1 == 'val_loss' or metric_1 == 'val_mean_loss':
                ES(EMS.dict_val_metric[metric_1][-1], None)
                if ES.early_stop == False:
                    # loss
                    if selected_EMS.dict_val_metric[metric_1] >= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            else:
                ES(None, EMS.dict_val_metric[metric_1][-1])
                if ES.early_stop == False:
                    # accuracy, AUC
                    if selected_EMS.dict_val_metric[metric_1] <= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')

        else:
            if selected_EMS.latest_selceted_model_dir != '':
                os.remove(selected_EMS.latest_selceted_model_dir)

            current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
            try:
                torch.save(model.state_dict(), current_model_dir)
            except KeyboardInterrupt:
                pass
            except ValueError:
                pass
            selected_EMS.latest_selceted_model_dir = current_model_dir
            tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')
    else:
        """save model"""
        if selected_EMS.latest_selceted_model_dir_2 != '':
            os.remove(selected_EMS.latest_selceted_model_dir_2)
        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
        try:
            torch.save(model.state_dict(), current_model_dir)
        except KeyboardInterrupt:
            pass
        except ValueError:
            pass
        selected_EMS.latest_selceted_model_dir_2 = current_model_dir
        tmp_flag = True
    return tmp_flag

class eval_selected_metirc_storage():
    def __init__(self):
        super(eval_selected_metirc_storage, self).__init__()

        """ saved model info"""
        self.latest_selceted_model_dir = ''
        self.latest_selceted_model_dir_2 = ''
        self.selected_ep = 0
        self.dict_val_metric = {
            'val_mean_loss': 10000,
            'val_loss': 10000,
            'val_acc': 0,
            'val_auc': 0,
        }

class eval_metric_storage():
    def __init__(self):
        super(eval_metric_storage, self).__init__()
        self.total_step = 0
        self.LR = []

        """ train """
        self.train_loss = []
        self.train_aux_loss_1 = []
        self.train_aux_loss_2 = []
        self.train_aux_loss_3 = []
        self.train_aux_loss_4 = []
        self.train_acc = []
        # self.train_auc = []
        self.train_step = []

        """ val """
        self.dict_val_metric = {
            'val_loss_queue': deque([]),
            'val_mean_loss': [],
            'val_loss': [],
            'val_acc': [],
            'val_auc': [],
            'val_MAE': [],
        }
        self.val_step = []

        """ test """
        self.test_loss = []
        self.test_acc = []
        self.test_auc = []
        self.test_MAE = []
        self.test_step = []

    def forward(self):
        pass



class EarlyStopping():
    """
    Early Stopping to terminate training early under certain conditions
    """
    def __init__(self, delta=0, patience=5, verbose = True):
        self.delta = delta
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_loss = None
        self.best_acc = None
        self.best_mean_loss = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.val_acc_max = 0
        self.wait = 0
        self.stopped_epoch = 0
        super(EarlyStopping, self).__init__()

    def __call__(self, val_loss, val_acc):
        if self.early_stop == False:
            if val_loss != None:
                if self.best_loss is None:
                    self.best_loss = val_loss
                # better model has been found.
                if val_loss < self.best_loss + self.delta:
                    self.best_loss = val_loss
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True
            else:
                if self.best_acc is None:
                    self.best_acc= val_acc
                # better model has been found.
                if val_acc > self.best_acc + self.delta:
                    self.best_acc = val_acc
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True

            if self.verbose == True:
                print(f'Early Stopping counter : {self.counter} out of {self.patience}')
        else:
            pass




def min_max_norm(img):
    max_val = img.max()
    min_val = img.min()
    if max_val==min_val:
        return img
    else:
        norm_img = (img-min_val)/(max_val-min_val)
        return norm_img


def Gauss_Norm_voxelWise(X, mu=0, sigma=1 , train = True):
    # print('Gaussian Normalize')
    if train == True:
        mu = np.expand_dims(np.mean(X, 0), axis=0)  # (1, 176, 210, 168)
        # TODO sigma != 0 , sigma != inf
        sigma = np.expand_dims(np.std(X, 0), axis=0) + np.spacing(1)  # (1, 176, 210, 168)

        norm_Data = np.zeros([X.shape[0],X.shape[1],X.shape[2],X.shape[3]])
        norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        if np.sum(np.isinf(norm_Data[:, :, :, :])) != 0:
            norm_Data[:, :, :, :][np.isinf(norm_Data[:, :, :, :])] = 0
            print("isinf!!!")
            assert np.sum(np.isinf(norm_Data[:, :, :, :])) != 0
        # for i in range(X.shape[0]):
        #     # print('train %d'%i)
        #     norm_Data[i,:,:,:] = np.nan_to_num((X[i,:,:,:]-mu) / sigma)
        #     if np.sum(np.isinf(norm_Data[i,:,:,:])) != 0 :
        #         norm_Data[i,:,:,:][np.isinf(norm_Data[i,:,:,:])] = 0
        #         print("isinf!!!")
        #         assert np.sum(np.isinf(norm_Data[i, :, :, :])) != 0

        return norm_Data, mu, sigma

    else:
        norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3]])
        norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        if np.sum(np.isinf(norm_Data[:, :, :, :])) != 0:
            norm_Data[:,:,:,:][np.isinf(norm_Data[:,:,:,:])] = 0
            print("isinf!!!")
            print(np.sum(np.isinf(norm_Data[:,:,:,:])))
            assert np.sum(np.isinf(norm_Data[:, :, :, :])) != 0
        # for i in range(X.shape[0]):
        #     # print('test %d' % i)
        #     norm_Data[i, :, :, :] = np.nan_to_num((X[i, :, :, :] - mu) / sigma)
        #     # print(np.sum(np.isinf(norm_Data[i,:,:,:])))
        #     if np.sum(np.isinf(norm_Data[i,:,:,:])) != 0:
        #         # norm_Data[i,:,:,:][np.isinf(norm_Data[i,:,:,:])] = 0
        #         print("isinf!!!")
        #         assert np.sum(np.isinf(norm_Data[i, :, :, :])) != 0
        return norm_Data


def Gauss_Norm_subjectWise(X, mu=0, sigma=1 , train = True):
    if train == True:
        mu = np.mean(X,  keepdims=True)# (1,1,1,1)
        sigma = np.std(X, keepdims=True) + np.spacing(1)# (1,1,1,1)

        # TODO sigma != 0 , sigma != inf
        assert sigma != 0
        assert sigma != np.inf
        # norm_Data = np.zeros([X.shape[0],X.shape[1],X.shape[2],X.shape[3]])
        # norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        # norm_Data = np.nan_to_num((X[:, :, :, :] - mu) / sigma)

        norm_Data = (X - mu) / sigma
        assert np.sum(np.isinf(norm_Data)) == 0
        # norm_Data[:,:,:,:][np.isinf(norm_Data[:,:,:,:])] = 0

        return norm_Data, mu, sigma

    else :
        # norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3]])
        # norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        # norm_Data = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        assert sigma != 0
        assert sigma != np.inf
        norm_Data = (X - mu) / sigma
        assert np.sum(np.isinf(norm_Data)) == 0

        return norm_Data


def data_normalization(X, min = 0 , max = 255):
    norm_Data = np.subtract(X, min)
    norm_Data = np.nan_to_num(np.divide(norm_Data, (max - min)/2))
    norm_Data = np.subtract(norm_Data , 1)
    return norm_Data

def data_minmax(X):
    shape = [None]*len(X)
    for i in range(len(X)):
        shape[i]=X.size()[i]
    for i in range(shape[0]):
        for j in range(shape[1]):
            X[i,j,:,:,:] = min_max_norm(X[i,j,:,:,:])

    return X

def data_mean_centering(X):
    norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3], X.shape[4]])
    for i in range(X.shape[0]):
        mu = np.mean(np.array(X[i]), keepdims=True)
        norm_Data[i] = np.subtract(X[i] , mu)
    return norm_Data


def Cross_validation(num_data, k_fold, Random_seed=0):
    indices = np.random.RandomState(seed=Random_seed).permutation(num_data)
    np.random.shuffle(indices)
    num_idx = num_data // k_fold
    sample_remainder = num_data % k_fold
    list_size_each_fold = []
    tmp = 0
    for i_fold in range(k_fold):
        if sample_remainder > i_fold:
            list_size_each_fold.append(num_idx + 1 + tmp)
        else:
            list_size_each_fold.append(num_idx+ tmp)
        tmp = list_size_each_fold[-1]

    train_idx = []
    test_idx = []
    val_idx = []
    for i_fold in range(k_fold):
        fold_slice = np.split(indices.copy(), list_size_each_fold, axis=0)
        fold_slice.pop(-1)
        if i_fold == k_fold - 1:
            test_idx.append(fold_slice.pop(i_fold % 10))
            val_idx.append(fold_slice.pop(0))
        else:
            test_idx.append(fold_slice.pop(i_fold % 10))
            val_idx.append(fold_slice.pop(i_fold % 10))
        train_idx.append(np.concatenate(fold_slice))
    return train_idx, val_idx, test_idx



def search_in_whole_subdir(file_dir, sub_dir, n_file, n_ext='.nii'):
    """
    :param file_dir: file directory
    :param sub_dir: the directory default = ''
    :param n_file: a list which words that extraction included
    :param n_ext: the type of files (e.g., .gt, .nii)
    :return: file list
    """

    """ make dir to save if not exist """
    if os.path.exists(file_dir + sub_dir) == False:
        os.makedirs(file_dir+sub_dir)

    file_list = [] # the list to reture
    for (path, dir, files) in os.walk(file_dir + sub_dir):
        # print(path)
        for filename in files:
            ext = os.path.splitext(filename)[-1] # 0 : filename, 1 : 확장자
            _file = os.path.splitext(filename)[0]
            if ext == n_ext:
                count_ = 0
                for i in range (len(n_file)):
                     if n_file[i] in _file :
                         count_ += 1
                if count_ == len(n_file) :
                    file_to_save = path + '/' + filename
                    file_list.append(file_to_save)
    # print(len(file_list))
    return file_list

def save_tensor_to_img(tensor, name):
    tmp_array = tensor.data.cpu().numpy()
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[0, 0, int(shape[2]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[0, 0, :, int(shape[3] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[0, 0, : , :, int(shape[4] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(st.save_plot_dir + name + ".png")
    plt.close(fig)


def plot_confusion_matrix(y_true, y_pred, classes,f_dir, f_name, normalize=False, title=None, cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    """
    if not title:
        if normalize:
            title = 'Normalized confusion matrix'
        else:
            title = 'Confusion matrix, without normalization'

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)

    # Only use the labels that appear in the data
    classes = classes[unique_labels(y_true, y_pred)]
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    print(cm)

    plt.rcParams.update({'font.size': 10})
    fig, ax = plt.subplots()
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
    ax.figure.colorbar(im, ax=ax)

    # We want to show all ticks...
    ax.set(xticks=np.arange(cm.shape[1]),
           yticks=np.arange(cm.shape[0]),
           # ... and label them with the respective list entries
           xticklabels=classes, yticklabels=classes,
           title=title,
           ylabel='True label',
           xlabel='Predicted label')

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor")

    # Loop over data dimensions and create text annotations.
    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], fmt),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black")
    fig.tight_layout()
    plt.savefig(f_dir + f_name)

def save_tensor_to_img(tensor, save_dir):
    tmp_array = tensor.data.cpu().numpy()
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[0, 0, int(shape[2]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[0, 0, :, int(shape[3] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[0, 0, : , :, int(shape[4] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(save_dir)
    plt.close(fig)

def save_numpy_to_2D_img(img, save_dir, file_name= 'test'):
    make_dir(save_dir)
    tmp_array = img
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[int(shape[0]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[:, int(shape[1] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[: , :, int(shape[2] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(save_dir + file_name)
    plt.close(fig)

def save_featureMap_tensor(tensor, dirToSave = './', name='test'):
    tmp_dir = dirToSave +'/featuremap'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)
    tmp_array = tensor.data.cpu().numpy()
    f_img = nib.Nifti1Image(tmp_array, np.eye(4))
    nib.save(f_img, os.path.join(tmp_dir + '/'+ name+ '.nii.gz'))

def save_featureMap_numpy(numpy, dirToSave = './', name='test'):
    tmp_dir = dirToSave +'/featuremap'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)
    f_img = nib.Nifti1Image(numpy, np.eye(4))
    nib.save(f_img, os.path.join(tmp_dir + '/'+ name+ '.nii.gz'))


def preparation_fold_index(config):
    list_trIdx = [] # (# of class, # of fold)
    list_valIdx = []  # (# of class, # of fold)
    list_teIdx = []  # (# of class, # of fold)

    for i in range(len(st.list_class_type)):
        if st.list_data_type[st.data_type_num] == 'Density':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float64).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float32).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_Jacob_256' or st.list_data_type[st.data_type_num] == 'ADNI_AAL_256':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.uint8).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        print(tmp_shape)
        tmp_trIdx, tmp_valIdx, tmp_teIdx = utils.Cross_validation(tmp_shape, config.kfold, Random_seed=0)
        list_trIdx.append(tmp_trIdx)
        list_valIdx.append(tmp_valIdx)
        list_teIdx.append(tmp_teIdx)

    """ Check whether all of the index is different """
    for i_class_type in range(len(st.list_class_type)):
        for i_fold in range(config.kfold):

            for j_class_type in range(len(st.list_class_type)):
                for j_fold in range(config.kfold):

                    if i_fold != j_fold  or i_class_type != j_class_type:
                        assert not(np.array_equal(list_trIdx[i_class_type][i_fold], list_trIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_valIdx[i_class_type][i_fold], list_valIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_teIdx[i_class_type][i_fold], list_teIdx[j_class_type][j_fold]))


    """ save index for each class """
    for i_class_type in range(len(st.list_class_type)):
        with open(st.train_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_trIdx[i_class_type], fp)
        with open(st.val_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_valIdx[i_class_type], fp)
        with open(st.test_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_teIdx[i_class_type], fp)

def plot_list_v1(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax'):
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)

    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10 * len(y)))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})

    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag == 'minmax':
        for i in range(len(y)):
            y_range[0] = np.vstack(y).min()
            y_range[1] = np.vstack(y).max()
    elif flag == 'acc':
        y_range[0] = 40
        y_range[1] = 100
    elif flag == 'loss':
        y_range[0] = 0
        y_range[1] = 4

    ##TODO: plotting ans save
    for i in range(len(y)):
        ax1 = fig.add_subplot(len(y), 1, i + 1)
        # ax1.set_title(title + '_{}'.format(i))
        ax1.set_ylabel(n_ylabel[i], color='b')
        ax1.set_xlabel(n_xlabel, color='b')
        ax1.plot(x, y[i], c='b', ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        plt.grid(True)
        plt.legend()
        np.save(file=save_dir + file_name + '_' + n_ylabel[i], arr=y[i])

    if (x_range[0] - x_range[1]) != 0:
        plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
    if (y_range[0] - y_range[1]) != 0:
        plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)

    plt.savefig(os.path.join(save_dir + file_name))

    plt.close('all')

def plot_list_v2(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax'):
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)

    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})
    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag == 'minmax':
        for i in range(len(y)):
            y_range[0] = np.vstack(y).min()
            y_range[1] = np.vstack(y).max()
    elif flag == 'acc':
        y_range[0] = 40
        y_range[1] = 100
    elif flag == 'loss':
        y_range[0] = 0
        y_range[1] = 4

    ##TODO: plotting ans save
    ax1 = fig.add_subplot(1, 1, 1)
    # ax1.set_title(title)
    # ax1.set_ylabel(n_ylabel, color='k')
    ax1.set_xlabel(n_xlabel, color='k')

    list_color = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
    for i in range(len(y)):
        ax1.plot(x, y[i], c=list_color[i], ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        np.save(file=save_dir + file_name +'_' + n_ylabel[i], arr=y[i])

    if (x_range[0] - x_range[1]) != 0:
        plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
    if (y_range[0] - y_range[1]) != 0:
        plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)

    plt.grid(True)
    plt.legend()
    plt.savefig(os.path.join(save_dir + file_name))
    plt.close('all')


def dfs_freeze(model, requires_grad = False):
    for name, child in model.named_children():
        for param in child.parameters():
            # print(child)
            param.requires_grad = requires_grad
        dfs_freeze(child)

def model_dir_to_load(fold, model_load_dir):
    """ find the maximum epoch model between saved models"""
    included_file_name = ['fold' + str(fold)]
    # get the model corresponding to the specific fold
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    s_index = 0 # start index
    e_index = 0 # end index
    n_epoch = []
    for i in range (len(models)):
        for j in range (len(models[i])):
            if models[i][-(j+1)] == 'h':
                s_index = j
                break

        for j in range (len(models[i])):
            if models[i][-(j+1)] == '.':
                e_index = j
                break
        n_epoch.append(models[i][-(s_index+1)+1 : -(e_index+1)])
    included_file_name.append(max(n_epoch))
    # get the model corresponding to the specific max epoch
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    model_dir = models[0]
    return model_dir

def model_dir_to_load_2(fold, model_load_dir):
    """ find the maximum epoch model between saved models"""
    included_file_name = ['fold' + str(fold)]
    # get the model corresponding to the specific fold
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    s_index = 0 # start index
    e_index = 0 # end index
    n_epoch = []
    for i in range (len(models)):
        for j in range (len(models[i])):
            if models[i][-(j+1)] == 'h':
                s_index = j
                break

        for j in range (len(models[i])):
            if models[i][-(j+1)] == '.':
                e_index = j
                break
        n_epoch.append(models[i][-(s_index+1)+1 : -(e_index+1)])
    included_file_name.append(max(n_epoch))
    # get the model corresponding to the specific max epoch
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    model_dir = models[0]
    return model_dir

def tensor_cropping(x, RF_size = 33) :
    """
    x =  tensor
    RF_size = receptive_filed_size
    """
    tmp_size = [None]*3
    for i in range(3):
        tmp_size[i] = x.size()[i+2]

    standard = RF_size // 2
    tmp_coord = np.random.randint(standard, size=3)
    out = x[:, :,
          tmp_coord[0]: tmp_size[0] - standard + tmp_coord[0],
          tmp_coord[1]: tmp_size[1] - standard + tmp_coord[1],
          tmp_coord[2]: tmp_size[2] - standard + tmp_coord[2],
          ]
    return out

def smooth_one_hot(true_labels, classes=2, smoothing=0.9):
    """
    if smoothing == 0, it's one-hot method
    if 0 < smoothing < 1, it's smooth method

    """
    assert 0 <= smoothing < 1
    confidence = 1.0 - smoothing
    label_shape = torch.Size((true_labels.size(0), classes))
    with torch.no_grad():
        true_dist = torch.empty(size=label_shape, device=true_labels.device)
        true_dist.fill_(smoothing / (classes - 1))
        # true_dist.scatter_(1, true_labels.data.unsqueeze(1), confidence)
        true_dist.scatter_(1, true_labels.data, confidence)
    return true_dist

def roll_tensor(x, n, axis):
    if axis == 0 :
        return torch.cat((x[-n:, :, :], x[:-n, :, ]), dim=0)
    elif axis == 1 :
        return torch.cat((x[:, -n:, :], x[:, :-n, ]), dim=1)
    elif axis == 2 :
        return torch.cat((x[:, :, -n:], x[:, :, :-n]), dim=2)

def push_tensor(x, n, axis):
    zero_tensor = torch.zeros_like(x)
    if axis == 0 :
        if n > 0 :
            return torch.cat((zero_tensor[-n:, :, :], x[:-n, :, :]), dim=0)
        elif n < 0 :
            return torch.cat((x[-n:, :, :], zero_tensor[:-n, :, :]), dim=0)
        else:
            return x
    elif axis == 1 :
        if n > 0:
            return torch.cat((zero_tensor[:, -n:, :], x[:, :-n, :]), dim=1)
        elif n < 0:
            return torch.cat((x[:, -n:, :], zero_tensor[:, :-n, :]), dim=1)
        else:
            return x

    elif axis == 2 :
        if n > 0:
            return torch.cat((zero_tensor[:, :, -n:], x[:, :, :-n]), dim=2)
        elif n < 0:
            return torch.cat((x[:, :, -n:], zero_tensor[:, :, :-n]), dim=2)
        else:
            return x


def crop_tensor(datas):
    list_cropping_info = [[] for tmp_i in range(3)]
    """ width """
    tmp_size = [a_i - b_i for a_i, b_i in zip(st.max_crop_size, st.min_crop_size)]
    width_size = [np.random.randint(k) + st.min_crop_size[i] for i, k in enumerate(tmp_size)]

    """ start """
    img_size = [st.x_size , st.y_size, st.z_size]
    start_point = [np.random.randint(img_size[i] - i_width) for i, i_width in enumerate(width_size)]

    for i in range(len(img_size)):
        list_cropping_info[i].append(start_point[i])
        list_cropping_info[i].append(width_size[i])

    return datas[:, :, list_cropping_info[0][0]:list_cropping_info[0][0] + list_cropping_info[0][1],
            list_cropping_info[1][0]:list_cropping_info[1][0] + list_cropping_info[1][1],
            list_cropping_info[2][0]:list_cropping_info[2][0] + list_cropping_info[2][1]], list_cropping_info




def excel_setting(start_fold, end_fold, result_dir, f_name):
    """ setting for the excel file """
    wb = Workbook()
    ws1 = wb.create_sheet('train_result', 0)
    exp_name = st.exp_name
    exp_description = st.exp_description

    """excel setting"""

    """ first col"""
    ws1.cell(row=1 + st.push_start_row, column=1, value="fold")
    for i in range(len(st.list_eval_metric)):
        ws1.cell(row=2+i + st.push_start_row, column=1, value=st.list_eval_metric[i])

    """ first row"""
    for col in range(start_fold, end_fold + 1):
        ws1.cell(row=1 + st.push_start_row, column=col + 1, value="fold_" + str(col))
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2, value="Avg")
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2).font = Font(name='Calibri', size=12, bold=True)
    column = str(chr(64 + end_fold + 2))
    ws1.column_dimensions[column].width = 20

    """ head """
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.merge_cells(start_row=1, end_row = 1, start_column= 1, end_column=n_col)
    ws1.merge_cells(start_row=2, end_row = 2, start_column= 1, end_column=n_col)
    ws1.cell(row=1, column=1, value=exp_name)
    ws1.cell(row=2, column=1, value=exp_description)


    box = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))

    """save xlsx"""
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.column_dimensions['A'].width = 20
    for i_row in range(1, n_row+1):
        for i_col in range(1, n_col+1):
            ca1 = ws1.cell(row = i_row, column = i_col)
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1.border = box
            if i_col == 1:
                ca1.font = Font(name='Calibri', size = 15, bold=True)
    wb.save(result_dir + "/{}.xlsx".format(f_name))
    return wb, ws1

def analysis_AttentionMap(fold, attn, batch_num, label, dim ):
    # ut.analysis_AttentionMap(fold=fold, attn=attn, batch_num=test_batchnum, label=labels.data.cpu().numpy().squeeze(), dim=(11, 14, 11))

    tmp = attn[0].data.cpu().numpy()
    tmp_dir = './attention_analysis'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)

    for i_head in range(tmp.shape[0]):
        std = np.std(tmp[i_head], axis=0)
        mean = np.mean(tmp[i_head], axis=0)
        x = np.arange(0, tmp.shape[-1], 1)

        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax_1 = fig.add_subplot(211)
        ax_2 = fig.add_subplot(212)

        # Turn off axis lines and ticks of the big subplot
        # ax.spines['top'].set_color('none')
        # ax.spines['bottom'].set_color('none')
        # ax.spines['left'].set_color('none')
        # ax.spines['right'].set_color('none')
        ax.tick_params(labelcolor='w', top=False, bottom=False, left=False, right=False)

        ax_1.plot(x, mean)
        ax_2.plot(x, std)

        # Set common labels
        ax.set_xlabel('position')
        # ax.set_ylabel('mean & std')

        ax_1.set_title('mean')
        ax_2.set_title('std')

        # plt.show()
        plt.savefig(tmp_dir + '/sum_std_of_attn_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))

        save_featureMap_numpy(mean.reshape(dim), './attention_mean', 'mean_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))
        save_featureMap_numpy(std.reshape(dim), './attention_std', 'std_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))

        # plt.imshow(tmp[i_head])
        # plt.colorbar()
        # plt.savefig(tmp_dir + '/a_a_matrix_head_{}_fold_{}'.format(i_head, fold))
        plt.close('all')
    return None

def make_dir(dir = './', flag_rm = False):
    if os.path.exists(dir) == True:
        if flag_rm == True:
            shutil.rmtree(dir)
            os.makedirs(dir)
    else:
        os.makedirs(dir)


def eval_classification_model(config,  loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    with torch.no_grad():
        for datas, labels, alabel, mlabel in loader:
            """ input"""
            datas = Variable(
                datas[:, :, :, :, :].view(-1, 1, config.modality, config.sagital, config.coronal, config.axial)).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """

            dict_result = model(datas)
            # dict_result = model(datas, alabel.cuda())
            output_1 = dict_result['logits']
            output_2 = dict_result['Aux_logits']
            output_3 = dict_result['logitMap']

            """ loss 1 """
            loss_list_1 = []
            if fst.flag_loss_1 == True:
                loss_2 = criterion(output_1, labels)
                loss_list_1.append(loss_2)

            ## TODO : aux_loss
            if fst.flag_loss_2 == True:
                loss_2 = criterion(output_2, labels)
                loss_list_1.append(loss_2)

            ## TODO : patch-level loss
            if fst.flag_loss_3 == True:
                tmp_shape = output_3.shape
                logits = output_3.view(tmp_shape[0], tmp_shape[1], -1)
                tmp_shape = logits.shape
                loss_list_2 = []
                for i_patch in range(tmp_shape[-1]):
                    loss_list_2.append(criterion(logits[:, :, i_patch], labels))
                loss_2 = sum(loss_list_2)/ len(loss_list_2)
                loss_list_1.append(loss_2)
            loss = sum(loss_list_1)

            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_1)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result

def eval_classification_model_esemble(config, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    with torch.no_grad():
        """ get the output, logit"""
        model.eval()

        for datas, labels, alabel, mlabel in loader:
            stride_for_assemble = st.size_translation // 2
            list_logits = []
            list_loss = []
            for i in range(2):
                for j in range(2):
                    for k in range(2):
                        """ input"""
                        datas = Variable(datas[:, :, i * stride_for_assemble:, j * stride_for_assemble:, k * stride_for_assemble:]).cuda()
                        labels = Variable(labels.long()).cuda()

                        """ forward propagation """
                        dict_result = model(datas)
                        # dict_result = model(datas, alabel.cuda())
                        output_logit = dict_result['logits']


                        """ calculate the loss """
                        loss = criterion(output_logit, labels)
                        list_logits.append(output_logit.data.cpu().numpy())
                        list_loss.append(loss.data.cpu().numpy())

            """ add the loss """
            # loss_np += loss.data.cpu().numpy()
            loss_np += np.mean(list_loss)
            f_output_logit = np.mean(np.vstack(list_logits), axis = 0, keepdims=True)

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(torch.tensor(f_output_logit).cuda())
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result

def eval_classification_model_cropped_input(config,  loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    tmp_size_x_1 = (st.x_size - st.max_crop_size) // 2
    tmp_size_x_2 = tmp_size_x_1 + st.max_crop_size

    tmp_size_y_1 = (st.y_size - st.max_crop_size) // 2
    tmp_size_y_2 = tmp_size_y_1 + st.max_crop_size

    tmp_size_z_1 = (st.z_size - st.max_crop_size) // 2
    tmp_size_z_2 = tmp_size_z_1 + st.max_crop_size
    with torch.no_grad():
        for datas, labels, alabel, mlabel in loader:
            """ input"""
            datas = Variable(
                datas[:, :, tmp_size_x_1:tmp_size_x_2, tmp_size_y_1:tmp_size_y_2, tmp_size_z_1:tmp_size_z_2]).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """

            dict_result = model(datas)
            # dict_result = model(datas, alabel.cuda())
            output_logit = dict_result['logits']

            """ calculate the loss """
            loss = criterion(output_logit, labels)

            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_logit)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


def eval_classification_using_pretrained(config,  loader, model, model_1, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    # relevance_map = [[], []]
    with torch.no_grad():
        for datas, labels, alabel, mlabel in loader:
            """ input"""
            datas = Variable(
                datas[:, :, :, :, :].view(-1, 1, config.modality, config.sagital, config.coronal, config.axial)).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            dict_result = model_1(datas)
            final_evidence = dict_result['final_evidence']
            featureMaps = dict_result['featureMaps']
            dict_result = model(datas, final_evidence, featureMaps)
            output_logit = dict_result['logits']

            """ calculate the loss """
            loss = criterion(output_logit, labels)

            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_logit)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    # relevance_map_0 = np.vstack(relevance_map[0])
    # ut.save_featureMap_numpy(np.mean(relevance_map_0, axis=0)[0], './', 'relevance_{}_logit_0'.format(st.list_selected_for_train[0]))
    # ut.save_featureMap_numpy(np.mean(relevance_map_0, axis=0)[1], './', 'relevance_{}_logit_1'.format(st.list_selected_for_train[0]))
    # relevance_map_1 = np.vstack(relevance_map[1])
    # ut.save_featureMap_numpy(np.mean(relevance_map_1, axis=0)[0], './', 'relevance_{}_logit_0'.format(st.list_selected_for_train[1]))
    # ut.save_featureMap_numpy(np.mean(relevance_map_1, axis=0)[1], './', 'relevance_{}_logit_1'.format(st.list_selected_for_train[1]))

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


def plot_heatmap_with_overlay(orig_img, heatmap_img, save_dir, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    shape = heatmap_img.shape
    list_interval = []
    for j in range(3):
        tmp_list = []
        for i in np.arange(20, 81, 4):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    axis_type = ['Sagittal', 'Coronal', 'Axial']

    fig = plt.figure(figsize=(list_interval[0].shape[0] * 2, len(axis_type) * 3))
    plt.rcParams.update({'font.size': 20})
    fig.suptitle(fig_title, fontsize=30)

    heights = [1] * len(axis_type)
    widths = [10] * (list_interval[0].shape[0])
    widths.append(10)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )

    cmap_orig = plt.get_cmap('Greys')

    # cmap_heatmap = plt.get_cmap('Reds')
    cmap_heatmap = plt.get_cmap('coolwarm')
    # cmap_heatmap = plt.get_cmap('bwr')

    # for orig
    orig_vmax = np.percentile(orig_img, 100 - percentile)
    orig_vmin = np.percentile(orig_img, percentile)
    print(orig_vmin, orig_vmax)

    vmax = np.percentile(heatmap_img, 100-percentile)
    vmin = np.percentile(heatmap_img, percentile)
    # vmax = heatmap_img.max()
    # vmin = heatmap_img.min()
    print(heatmap_img.max())
    print(heatmap_img.min())
    print(vmin, vmax)

    if np.abs(vmax) > np.abs(vmin):
        vmax = np.abs(vmax)
        vmin = -np.abs(vmax)
    else:
        vmax = np.abs(vmin)
        vmin = -np.abs(vmin)

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh
    # thresh_max = np.percentile(heatmap_img, 97)
    # thresh_min = np.percentile(heatmap_img, 3)
    # print(thresh_min, thresh_max)
    # if np.abs(thresh_max) < np.abs(thresh_min):
    #     thresh_max = np.abs(thresh_max)
    #     thresh_min = -np.abs(thresh_max)
    # else:
    #     thresh_max = np.abs(thresh_min)
    #     thresh_min = -np.abs(thresh_min)

    alpha = 0.5
    axes = []
    for j, q in enumerate(axis_type):
        for i, p in enumerate(list_interval[j]):

            ax1 = fig.add_subplot(gs[j, i])

            if j == 0:
                orig_scattering_img = np.asarray(orig_img[int(p), :, :])
                heatmap_scattering_img = np.asarray(heatmap_img[int(p), :, :])
            elif j == 1:
                orig_scattering_img = np.asarray(orig_img[:, int(p), :])
                heatmap_scattering_img = np.asarray(heatmap_img[:, int(p), :])
            elif j == 2:
                orig_scattering_img = np.asarray(orig_img[:, :, int(p)])
                heatmap_scattering_img = np.asarray(heatmap_img[:, :, int(p)])

            orig_scattering_img = np.rot90(orig_scattering_img)
            heatmap_scattering_img = np.rot90(heatmap_scattering_img)
            heatmap_scattering_img[
                (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan

            if i == 0:
                # ax1.set_title(axis_type[j])
                ax1.set_ylabel(axis_type[j])
                # plt.ylabel(axis_type[j])
            ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
            # im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=positive_vmin, vmax=positive_vmax)
            im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
            ax1.set_yticks([])
            ax1.set_xticks([])
            ax1.spines['right'].set_visible(False)
            ax1.spines['top'].set_visible(False)
            ax1.spines['bottom'].set_visible(False)
            ax1.spines['left'].set_visible(False)
            axes.append(ax1)
            # ax1.axis('off')
            del orig_scattering_img, heatmap_scattering_img

    # (left, bottom, width, height)
    cax = plt.axes([0.95, 0.1, 0.01, 0.8])
    cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)

    cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
    cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
    # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

    plt.tight_layout()
    plt.savefig(save_dir)
    plt.close('all')


def data_augmentation(datas):
    list_cropping_info = None
    if fst.flag_translation == True:
        flag_rand_trans = np.random.randint(1000)
        if fst.flag_translation_ratio != 0:
            if flag_rand_trans % fst.flag_translation_ratio == 0:
                pass
            else:
                size_of_translation = st.size_translation
                translation_list = np.random.randint(size_of_translation, size=(datas.size()[0], 3)) - size_of_translation // 2
                for batch_i in range(translation_list.shape[0]):
                    for axis_i in range(translation_list.shape[1]):
                        datas[batch_i][0] = ut.push_tensor(datas[batch_i][0], translation_list[batch_i][axis_i], axis_i)
        else:
            size_of_translation = st.size_translation
            translation_list = np.random.randint(size_of_translation, size=(datas.size()[0], 3)) - size_of_translation // 2
            for batch_i in range(translation_list.shape[0]):
                for axis_i in range(translation_list.shape[1]):
                    datas[batch_i][0] = ut.push_tensor(datas[batch_i][0], translation_list[batch_i][axis_i], axis_i)

    if fst.flag_cropping == True:
        flag_rand_crop = np.random.randint(1000)
        if fst.flag_crop_ratio != 0:
            if flag_rand_crop % fst.flag_crop_ratio == 0:
                pass
            else:
                datas, list_cropping_info = ut.crop_tensor(datas)
        else:
            datas, list_cropping_info= ut.crop_tensor(datas)

    dict_result = {
        "datas": datas,
        "list_cropping_info": list_cropping_info,
    }
    return dict_result

def plot_training_info_1(fold, dir_pyplot, EMS):
    """ plot the chat"""
    """ train loss """
    y_list = []
    y_list_name = []
    y_list.append(EMS.train_loss)
    y_list_name.append('train loss')
    ut.plot_list_v1(EMS.train_step, y_list, title='train loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_train_loss'.format(fold), flag='minmax')
    del y_list, y_list_name

    """ train aux loss """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.train_aux_loss_1) == len(EMS.train_step):
        y_list.append(EMS.train_aux_loss_1)
        y_list_name.append('aux loss 1')
        tmp_flag = True
    if len(EMS.train_aux_loss_2) == len(EMS.train_step):
        y_list.append(EMS.train_aux_loss_2)
        y_list_name.append('aux loss 2')
        tmp_flag = True
    if len(EMS.train_aux_loss_3) == len(EMS.train_step):
        y_list.append(EMS.train_aux_loss_3)
        y_list_name.append('aux loss 3')
        tmp_flag = True
    if len(EMS.train_aux_loss_4) == len(EMS.train_step):
        y_list.append(EMS.train_aux_loss_4)
        y_list_name.append('aux loss 4')
        tmp_flag = True

    if tmp_flag == True:
        ut.plot_list_v1(EMS.train_step, y_list, title='train aux loss', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_train_aux_loss'.format(fold), flag='minmax')
    del y_list, y_list_name

    """ val test acc """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_acc']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_acc'])
        y_list_name.append('val acc')
        tmp_flag = True
    if len(EMS.test_acc) == len(EMS.val_step):
        y_list.append(EMS.test_acc)
        y_list_name.append('test acc')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='val_test_acc', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_val_test_acc'.format(fold), flag='minmax')
    del y_list, y_list_name

    """ val test loss """
    y_list = []
    y_list_name = []
    y_list.append(EMS.dict_val_metric['val_loss'])
    y_list_name.append('val loss')
    y_list.append(EMS.test_loss)
    y_list_name.append('test loss')
    ut.plot_list_v2(EMS.val_step, y_list, title='val_test_loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_val_test_loss'.format(fold), flag='minmax')
    del y_list, y_list_name

    """ learning rate """
    y_list = []
    y_list_name = []
    y_list.append(EMS.LR)
    y_list_name.append('learning rate')
    ut.plot_list_v1(EMS.val_step, y_list, title='Learning rate', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_Learning_rate'.format(fold), flag='minmax')
    del y_list, y_list_name

    """ train acc """
    if fst.flag_print_trainAcc == True:
        y_list = []
        y_list_name = ['train acc']
        y_list.append(EMS.train_acc)
        ut.plot_list_v1(EMS.train_step, y_list, title='train acc', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_train_acc'.format(fold), flag='minmax')
        del y_list, y_list_name


def eval_regression_model(config, loader, model, criterion):
    """ loader"""
    MAE_loss = 0
    criterion_MSE = nn.MSELoss()
    model.eval()
    torch.cuda.empty_cache()
    predict_result = []
    groundTruth=[]
    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            """ input"""
            datas = Variable(
                datas[:, :, :, :, :].view(-1, 1, config.modality, config.sagital, config.coronal, config.axial)).cuda()
            alabels = Variable(alabels.float()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            dict_result = model(datas)
            output_regression= dict_result['preds']
            predict_result.append(output_regression.data.cpu().numpy().squeeze())
            groundTruth.append(alabels.data.cpu().numpy().squeeze())

            """ calculate the loss """
            loss_1 = criterion(output_regression.squeeze(), alabels.squeeze())
            loss = loss_1

            MAE_loss += loss.data.cpu().numpy()
    total = len(loader.dataset)
    MAE_loss = MAE_loss / total
    predict_result_np = np.hstack(predict_result)
    groundTruth_np = np.hstack(groundTruth)
    r2 = r2_score(y_true=groundTruth_np, y_pred=predict_result_np )
    rmse = math.sqrt(mean_squared_error(y_true=groundTruth_np, y_pred=predict_result_np))
    dict_result = {
        "Loss": MAE_loss,
        "MAE": MAE_loss,
        "RMSE": rmse,
        "R_squared": r2,
    }
    return dict_result


def estimate_biological_age(age, MMSE, lambda_disease_factor):
    """ params """
    age_normalization_factor = 1 / 70

    if st.list_age_estimating_function[st.selected_function] == 'linear_1':
        # linear function
        age = age + (30 - MMSE) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'linear_1_with_age':
        # linear function with age
        age = age + (30 - MMSE) * lambda_disease_factor * age * age_normalization_factor

    elif st.list_age_estimating_function[st.selected_function] == 'sigmoid_1':
        # using sigmoid
        w_x = -0.1
        b_x = 15
        w_y = 30
        b_y = -5
        age = age + ((1 / (1 + np.exp(-(w_x * (MMSE - b_x))))) * w_y + b_y) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'sigmoid_1_with_age':
        # using sigmoid
        w_x = -0.1
        b_x = 15
        w_y = 30
        b_y = -5
        age = age + ((1 / (1 + np.exp(-(w_x * (MMSE - b_x))))) * w_y + b_y) * age * age_normalization_factor * lambda_disease_factor
    elif st.list_age_estimating_function[st.selected_function] == 'sqrt_1':
        # using sigmoid
        w = 4.5
        b = 30
        age = age + (w * (np.sqrt(-MMSE + b))) * lambda_disease_factor
    elif st.list_age_estimating_function[st.selected_function] == 'quadratic_1':
        # using sigmoid
        w = 0.05
        b = 30
        age = age + (w * np.power(MMSE-b, 2)) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'none':
        pass

    elif st.list_age_estimating_function[st.selected_function] == 'constant':
        age = age + 3 * lambda_disease_factor


    else:
        assert False, "the age estimation function has not been applied!"

    return age
